# Auto-generated from DEEL 2 notebook (handout generator)
from __future__ import annotations

PLACEHOLDER_MAPPING = {
  "L_0": [
    "L",
    "Geen"
  ],
  "L_2": [
    "L",
    "2 kolom"
  ],
  "L_3": [
    "L",
    "3 kolom"
  ],
  "L_4": [
    "L",
    "4 kolom"
  ],
  "L_5": [
    "L",
    "5 kolom"
  ],
  "L_6": [
    "L",
    "6 kolom"
  ],
  "L_7": [
    "L",
    "7 kolom"
  ],
  "M_lk_0": [
    "M (lk)",
    "Geen"
  ],
  "M_lk_2": [
    "M (lk)",
    "2 kolom"
  ],
  "M_lk_3": [
    "M (lk)",
    "3 kolom"
  ],
  "M_lk_4": [
    "M (lk)",
    "4 kolom"
  ],
  "M_lk_5": [
    "M (lk)",
    "5 kolom"
  ],
  "M_nws_0": [
    "M (nws)",
    "Geen"
  ],
  "M_nws_2": [
    "M (nws)",
    "2 kolom"
  ],
  "M_nws_3": [
    "M (nws)",
    "3 kolom"
  ],
  "M_nws_4": [
    "M (nws)",
    "4 kolom"
  ],
  "M_nws_5": [
    "M (nws)",
    "5 kolom"
  ],
  "S_lk_0": [
    "S (lk)",
    "Geen"
  ],
  "S_lk_2": [
    "S (lk)",
    "2 kolom"
  ],
  "S_lk_4": [
    "S (lk)",
    "4 kolom"
  ],
  "S_nws_0": [
    "S (nws)",
    "Geen"
  ],
  "S_nws_2": [
    "S (nws)",
    "2 kolom"
  ],
  "S_nws_4": [
    "S (nws)",
    "4 kolom"
  ],
  "XL_3": [
    "XL",
    "3 kolom"
  ],
  "XL_4": [
    "XL",
    "4 kolom"
  ],
  "XL_4B": [
    "XL",
    "4 kolom en bijplaat"
  ],
  "XL_5B": [
    "XL",
    "5 kolom en bijplaat"
  ],
  "XL_6": [
    "XL",
    "6 kolom"
  ],
  "XL_7": [
    "XL",
    "7 kolom"
  ],
  "XL_7B": [
    "XL",
    "7 kolom en bijplaat"
  ],
  "XS_0": [
    "XS",
    "Geen"
  ],
  "XS_4": [
    "XS",
    "4 kolom"
  ],
  "XXL_4": [
    "XXL",
    "4 kolom"
  ],
  "XXL_5B": [
    "XXL",
    "5 kolom en bijplaat"
  ],
  "XXL_6": [
    "XXL",
    "6 kolom"
  ],
  "XXL_6B": [
    "XXL",
    "6 kolom en bijplaat"
  ],
  "XXL_7": [
    "XXL",
    "7 kolom"
  ]
}

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether, ListFlowable, ListItem
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import pandas as pd, re, os

OPEN_SLOT_TEXT = "NOG ARTIKEL VOOR DEZE PLEK ZOEKEN"
CODE_RE = re.compile(r"[A-Z]\d{3}[A-Z]")

def generate_handout_pdf(xlsx_path, template_dir, logo_path, out_pdf=None):
    if out_pdf is None:
        out_pdf = os.path.splitext(os.path.basename(xlsx_path))[0] + "_handout.pdf"

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = set(wb.sheetnames)

    def get_log_ad(plaatsing):
        if plaatsing in sheet_names and wb[plaatsing]["AD1"].value:
            return str(wb[plaatsing]["AD1"].value)
        return ""

    def get_position_name_ae2(plaatsing):
        if plaatsing in sheet_names and wb[plaatsing]["AE2"].value:
            return str(wb[plaatsing]["AE2"].value)
        return ""

    def find_open_slot_placeholder(plaatsing):
        if plaatsing not in sheet_names:
            return None
        ws = wb[plaatsing]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and OPEN_SLOT_TEXT in cell.value.upper():
                    r = cell.row
                    return ws[f"X{r}"].value
        return None

    df = pd.read_excel(xlsx_path, sheet_name="Planning print")

    ph = df["Gekozen placeholder"].astype(str).str.strip()
    df["Artikel"] = ph.map(lambda x: PLACEHOLDER_MAPPING.get(x, ("",""))[0])
    df["Beeld (placeholder)"] = ph.map(lambda x: PLACEHOLDER_MAPPING.get(x, ("",""))[1])

    styles = getSampleStyleSheet()
    header_main = ParagraphStyle("HeaderMain", parent=styles["Heading2"], fontSize=18, leading=21, fontName="Helvetica-Bold")
    header_sub  = ParagraphStyle("HeaderSub", parent=styles["BodyText"], fontSize=10, leading=12, spaceAfter=6)
    section_style = ParagraphStyle("SectionTitle", parent=styles["Heading4"], fontSize=10.5, leading=12, spaceBefore=6, spaceAfter=2)
    cell_style = ParagraphStyle("Cell", parent=styles["BodyText"], fontSize=8.5, leading=10)
    header_cell_style = ParagraphStyle("HeaderCell", parent=styles["BodyText"], fontSize=9, leading=10)
    bullet_style = ParagraphStyle("Bullet", parent=styles["BodyText"], fontSize=8.5, leading=10, leftIndent=12)

    def esc(s):
        return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    def pcell(t, st=cell_style):
        return Paragraph(esc("" if pd.isna(t) else t), st)

    def article_word(n):
        return "artikel" if n == 1 else "artikelen"

    def extract_codes(t):
        if t is None or (isinstance(t, float) and pd.isna(t)):
            return []
        seen=set(); out=[]
        for c in CODE_RE.findall(str(t)):
            if c not in seen:
                seen.add(c); out.append(c)
        return out

    def find_img(code):
        pth = os.path.join(template_dir, f"{code}.jpg")
        return pth if os.path.exists(pth) else None

    def trunc_title(t, n=30):
        s = "" if pd.isna(t) else str(t)
        return s if len(s) <= n else s[:n] + "..."

    log_rules = [
        ("CM02", "<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W32 (104x70 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
        ("CM03", "<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W23 (158x94 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
        ("CM04", "<b>Vormgever:</b> Als noodgreep is een stopper-advertentie ingepland van het formaat W16 (266x94 mm). Plaats die eigen uiting of - beter nog - probeer die stopper overbodig te maken door bijvoorbeeld een foto bij een artikel groter te maken."),
        ("CI01", "<b>Samensteller:</b> Als noodgreep is een artikel S (zonder beeld) open gelaten."),
        ("CH01", "<b>Samensteller:</b> Als noodgreep is een artikel XS (zonder beeld) open gelaten."),
        ("CJ01", "<b>Samensteller:</b> Als absolute noodgreep is een fors artikel (M over groter) opengelaten. Bekijk wat er de afgelopen dagen is blijven liggen of overleg met de chef hoe deze ruimte kan worden gevuld."),
    ]

    logo_reader = ImageReader(logo_path)
    lw, lh = logo_reader.getSize()
    logo_aspect = lh / float(lw)

    header_h = 14*mm
    logo_h = 5.5*mm
    logo_w = logo_h / logo_aspect
    gap = 3*mm

    doc = SimpleDocTemplate(
        out_pdf,
        pagesize=landscape(A4),
        leftMargin=12*mm,
        rightMargin=12*mm,
        topMargin=header_h + 6*mm,
        bottomMargin=10*mm,
    )

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []
        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            total = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_header(total)
                super().showPage()
            super().save()
        def draw_header(self, total):
            left = doc.leftMargin
            right = doc.rightMargin
            pw, phh = self._pagesize
            top_y = phh - 6*mm

            self.drawImage(logo_reader, left, top_y-logo_h, width=logo_w, height=logo_h, mask="auto")
            self.setFont("Helvetica-Bold", 11)
            self.drawString(left + logo_w + gap, top_y - (logo_h*0.75), "Planning voor print")
            self.setFont("Helvetica", 9)
            self.drawRightString(pw-right, top_y - (logo_h*0.75), f"Pagina {self.getPageNumber()} van {total}")

            line_y = phh - header_h
            self.setStrokeColor(colors.lightgrey)
            self.setLineWidth(0.6)
            self.line(left, line_y, pw-right, line_y)

    headers = ["Artikel", "Beeld", "Titel", "Auteur", "Focusregio", "Top 8"]
    cols = ["Artikel", "Beeld (placeholder)", "Naam productie", "Auteur", "Focusregio", "Top 8"]

    dfg = df[["Plaatsing", "Gekozen template", "Gekozen placeholder", "Beeld voor print"] + cols].copy()

    total_w = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    col_widths = [0.07*total_w, 0.12*total_w, 0.41*total_w, 0.17*total_w, 0.12*total_w, 0.11*total_w]

    thumb_h, thumb_w = 78*mm, 180*mm

    elements = []

    for plaatsing, g in dfg.groupby("Plaatsing", sort=True):
        n = len(g)

        codes = []
        for t in g["Gekozen template"]:
            codes.extend(extract_codes(t))
        seen=set(); codes_u=[]
        for c in codes:
            if c not in seen:
                seen.add(c); codes_u.append(c)

        raw_templates = [t for t in g["Gekozen template"].dropna().astype(str).unique() if t.strip() != ""]
        if codes_u:
            template_label = " / ".join(codes_u[:2]) + (" …" if len(codes_u) > 2 else "")
        elif raw_templates:
            template_label = raw_templates[0] if len(raw_templates) == 1 else raw_templates[0] + " …"
        else:
            template_label = "—"

        pos_name = get_position_name_ae2(str(plaatsing)).strip()
        main_title = pos_name if pos_name else str(plaatsing)
        sub_title = f"({plaatsing}) ({n} {article_word(n)}, template {template_label})"

        block = [Paragraph(main_title, header_main), Paragraph(sub_title, header_sub)]

        show = codes_u[:2]
        if show:
            imgs = []
            for c in show:
                ip = find_img(c)
                if ip:
                    im = Image(ip)
                    im._restrictSize(thumb_w, thumb_h)
                    imgs.append(im)
            if imgs:
                prev = Table([imgs], colWidths=[thumb_w]*len(imgs), hAlign="LEFT")
                prev.setStyle(TableStyle([("BOX", (0,0), (-1,-1), 0.6, colors.lightgrey)]))
                block += [prev, Spacer(1, 6)]

        g_tbl = g.copy()
        log_str = get_log_ad(str(plaatsing))
        has_open_code = any(code in log_str for code in ["CI01", "CH01", "CJ01"])
        already_has_open = g_tbl["Naam productie"].astype(str).str.contains(OPEN_SLOT_TEXT, case=False, na=False).any()

        if has_open_code and not already_has_open:
            ph_open = find_open_slot_placeholder(str(plaatsing))
            if ph_open is not None:
                ph_open = str(ph_open).strip()
                art_open, beeld_open = PLACEHOLDER_MAPPING.get(ph_open, ("",""))
            else:
                art_open, beeld_open = "", ""

            g_tbl = pd.concat([g_tbl, pd.DataFrame([{
                "Plaatsing": plaatsing,
                "Gekozen template": "",
                "Gekozen placeholder": "",
                "Beeld voor print": "",
                "Artikel": art_open,
                "Beeld (placeholder)": beeld_open,
                "Naam productie": OPEN_SLOT_TEXT,
                "Auteur": "",
                "Focusregio": "",
                "Top 8": ""
            }])], ignore_index=True)

        data = [[pcell(h, header_cell_style) for h in headers]]
        for _, r in g_tbl.iterrows():
            data.append([pcell(r[c]) for c in cols])

        tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("ALIGN", (0,1), (1,-1), "CENTER"),
            ("ALIGN", (4,1), (5,-1), "CENTER"),
        ]))
        block.append(tbl)

        points = []

        if len(codes_u) == 2:
            points.append("<b>Vormgever:</b> Niet gekozen voor een spreadtemplate, maar voor twee enkele templates. Mogelijk staan beeld en koppen ongelukkig naast elkaar. Doe indien nodig aanpassing.")

        if any("variant" in str(t).lower() for t in raw_templates):
            points.append("<b>Vormgever en samensteller:</b> Een of meerdere vormen op de template moeten een handmatige reshape krijgen van nieuws naar lichte kop of omgekeerd. Het betreft daarbij per definitie een vorm van maat S of M.")

        for code, text in log_rules:
            if code in log_str:
                points.append(text)

        for _, r in g.iterrows():
            phs = "" if pd.isna(r["Gekozen placeholder"]) else str(r["Gekozen placeholder"])
            bvp = "" if pd.isna(r["Beeld voor print"]) else str(r["Beeld voor print"]).strip()
            tshort = trunc_title(r["Naam productie"], 30)

            if ("B" in phs) and (bvp != "Dragend en bijplaat"):
                points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{esc(tshort)}' is door de chef geen bijplaat gevraagd, maar in de planning kwam het wel goed uit om die toe te kennen. Beoordeel of er inderdaad nog een tweede beeld bij kan. Zo niet, bouw dan de vorm van dit verhaal enigszins om.")

            if (bvp == "Dragend en bijplaat") and ("B" not in phs):
                points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{esc(tshort)}' is door de chef een bijplaat gevraagd naast het dragende beeld, maar deze bijplaat kon bij de planning niet toegekend worden. Beoordeel of dit problematisch is en bouw de vorm van dit verhaal indien nodig enigszins om.")

            if phs.endswith("0") and (bvp not in ["", "Ongeschikt", "Flexibel"]):
                points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{esc(tshort)}' is door de chef ook Beeld gevraagd, maar dit kon bij de planning niet toegekend worden. Grijp alleen in als dit echt problematisch is.")

            if (not phs.endswith("0")) and (bvp == "Ongeschikt"):
                points.append(f"<b>Vormgever en samensteller:</b> Bij artikel '{esc(tshort)}' is door de chef aangegeven dat het beeld niet zo geschikt is voor print, maar als absolute noodgreep is er bij de planning toch voor gekozen om bij dit artikel een kleine plaat te gebruiken. In de uitzonderlijke situatie dat het online-beeld bij het verhaal echt niet kan voor print, los je het op met een stopper-advertentie.")

        if points:
            block += [
                Spacer(1, 6),
                Paragraph("Aandachtspunten", section_style),
                ListFlowable([ListItem(Paragraph(pt, bullet_style)) for pt in points], bulletType="bullet")
            ]

        block.append(Spacer(1, 14))
        elements.append(KeepTogether(block))

    doc.build(elements, canvasmaker=NumberedCanvas)
    return out_pdf

__all__ = ['generate_handout_pdf']
