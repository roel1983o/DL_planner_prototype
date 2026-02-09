# Auto-generated from PARSER notebook

import os
import io
import math
import re
import pandas as pd

UPLOAD_STATE = {
    "uploaded_filename": None,
    "output_filename": None,
    "output_df": None,
    "kandidaten_filename": None,
}

COLUMNS_TO_DROP = [
    "Story ID", "Story Status", "Slug",
    "Event date start", "Event time start",
    "Event date end", "Event time end",
    "Location", "Address", "ZIP code", "State", "City", "Country",
    "Vraag/Stelling Rhetoric",
    "Article DNA", "Platform", "Category", "Sub-category", "Type",
    "Scheduled publ. date", "Scheduled publ. time", "Scheduled time slot",
    "Actual publ. date", "Actual publ. time",
    "Scope", "EMBARGO", "Task",
    "Story creator - last name", "Story creator - first name",
    "Costs", "Currency", "Task note", "Asset Link"]

FINAL_COLUMNS_TO_DROP = COLUMNS_TO_DROP + [
    "Assignee last name",
    "Assignee first name",
    "Type verhaal",
]

GROUP_TO_LEVERANCIER = {
    "Midden-Limburg": "rMI",
    "Noord-Limburg": "rNO",
    "Maastricht - Heuvelland": "rMH",
    "Maastricht-Heuvelland": "rMH",
    "Sittard-Geleen": "rSG",
    "Parkstad": "rPS",
    "Nieuwsdienst": "rND",
    "Economie": "rEC",
    "Onderzoek": "rLS",
    "LS": "rLS",
}


# --- Placeholder mappingregels (uit Mappingregels_parser_new.xlsx) ---
PLACEHOLDER_RULES = [{'Artikelsoort': 'XXL',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'XXL_7',
  'Gewenste placeholder': 'XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'XXL_7',
  'Gewenste placeholder': 'XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'XXL_7, XXL_5B; XXL_6B',
  'Gewenste placeholder': 'XXL_4',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_6; XXL_5;',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'XXL_7, XXL_5B; XXL_6B',
  'Gewenste placeholder': 'XXL_4',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_6; XXL_5;',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_7; XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_7; XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'XXL_7',
  'Gewenste placeholder': 'XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'XXL_7',
  'Gewenste placeholder': 'XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': 'XXL_4',
  'Gewenste placeholder': 'XXL_5B; XXL_6B',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_7; XXL_6; XXL_5',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': 'XXL_4',
  'Gewenste placeholder': 'XXL_5B; XXL_6B',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_7; XXL_6; XXL_5',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_7; XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_7; XXL_6; XXL_5; XXL_4;',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XXL_5B; XXL_6B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_4',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XXL',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XXL_4',
  'Placeholder bij enigszins geschikt': 'XXL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'XL_7',
  'Gewenste placeholder': 'XL_3; XL_4; XL_5;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_4B; XL_5B; XL_6',
  'Vierde keus placeholder': 'XL_7B'},
 {'Artikelsoort': 'XL',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'XL_7',
  'Gewenste placeholder': 'XL_3; XL_4; XL_5;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_4B; XL_5B; XL_6',
  'Vierde keus placeholder': 'XL_7B'},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'XL_7B',
  'Gewenste placeholder': 'XL_3; XL_4;',
  'Placeholder bij enigszins geschikt': 'XL_3;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_5; XL_4B; XL_5B; XL_6',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'XL_7B',
  'Gewenste placeholder': 'XL_3; XL_4;',
  'Placeholder bij enigszins geschikt': 'XL_3;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_5; XL_4B; XL_5B; XL_6',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': 'XL_7B',
  'Gewenste placeholder': 'XL_4; XL_5; XL_6; XL_7',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_3; XL_4B; XL_5B;',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XL_3; XL_4; XL_5; XL_6;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_7; XL_4B; XL_5B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'XL_7B',
  'Gewenste placeholder': 'XL_3; XL_4; XL_6;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_4B; XL_5B; XL_7',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'XL_7B',
  'Gewenste placeholder': 'XL_3; XL_4; XL_6;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_4B; XL_5B; XL_7',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': 'XL_3',
  'Gewenste placeholder': 'XL_4B; XL_5B; XL_7B',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_4; XL_5; XL_6; XL_7',
  'Vierde keus placeholder': 'XXL_4; XXL_6'},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XL_4B; XL_5B; XL_7B',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_3; XL_4; XL_5; XL_6; XL_7',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': 'XL_7B; XL_7;',
  'Gewenste placeholder': 'XL_3; XL_4; XL_5; XL_6;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'XL_4B; XL_5B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': 'XL_7B; XL_7;',
  'Gewenste placeholder': 'XL_3; XL_4; XL_5; XL_6;',
  'Placeholder bij enigszins geschikt': 'XL_3; XL_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'XL_4B; XL_5B',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XL_3',
  'Placeholder bij enigszins geschikt': 'XL_3;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'XL_4;'},
 {'Artikelsoort': 'XL',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XL_3',
  'Placeholder bij enigszins geschikt': 'XL_3;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'XL_4'},
 {'Artikelsoort': 'L',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_0; L_2; L_3;',
  'Placeholder bij enigszins geschikt': 'L_0; L_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'L_4; L_5'},
 {'Artikelsoort': 'L',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'L_3',
  'Gewenste placeholder': 'L_0; L_2',
  'Placeholder bij enigszins geschikt': 'L_0; L_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'L_4'},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_2; L_3; L_4',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'L_0',
  'Vierde keus placeholder': 'L_5'},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'L_4',
  'Gewenste placeholder': 'L_2; L_3;',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'L_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_4; L_5; L_6; L_7',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'L_2; L_3',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_3; L_4; L_5; L_6',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'L_2; L_7',
  'Vierde keus placeholder': 'L_0'},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_2; L_3; L_4; L_5; L_6',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'L_7',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'L_0',
  'Gewenste placeholder': 'L_2; L_3; L_4; L_5;',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'L_6; L_7',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_5; L_6; L_7',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'L_2; L_3; L_4',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'L_5; L_6; L_7',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'L_2; L_3; L_4',
  'Vierde keus placeholder': 'L_0'},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': 'L_7',
  'Gewenste placeholder': 'L_0; L_2; L_3; L_4; L_5;',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'L_6',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': 'L_7',
  'Gewenste placeholder': 'L_0; L_2; L_3; L_4; L_5;',
  'Placeholder bij enigszins geschikt': 'L_0; L_2; L_3; L_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'L_6',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'L_2',
  'Gewenste placeholder': 'L_0',
  'Placeholder bij enigszins geschikt': 'L_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'L_3'},
 {'Artikelsoort': 'L',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'L_2',
  'Gewenste placeholder': 'L_0',
  'Placeholder bij enigszins geschikt': 'L_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_0; M_lk_2; M_lk_3',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'M_lk_4'},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_0; M_lk_2;',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'M_lk_4',
  'Gewenste placeholder': 'M_lk_2; M_lk_3;',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_lk_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'M_lk_4',
  'Gewenste placeholder': 'M_lk_2; M_lk_3;',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_lk_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': 'M_lk_2',
  'Gewenste placeholder': 'M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_lk_3;',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_3; M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_lk_2',
  'Vierde keus placeholder': 'M_lk_0'},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_2; M_lk_3; M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'M_lk_0',
  'Gewenste placeholder': 'M_lk_2; M_lk_3; M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_lk_3',
  'Vierde keus placeholder': 'M_lk_2'},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_lk_3',
  'Vierde keus placeholder': 'M_lk_2'},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_0; M_lk_2; M_lk_3; M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_lk_0; M_lk_2; M_lk_3; M_lk_4; M_lk_5',
  'Placeholder bij enigszins geschikt': 'M_lk_0; M_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'M_lk_2',
  'Gewenste placeholder': 'M_lk_0',
  'Placeholder bij enigszins geschikt': 'M_lk_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'M_lk_3'},
 {'Artikelsoort': 'M_lk',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'M_lk_2',
  'Gewenste placeholder': 'M_lk_0',
  'Placeholder bij enigszins geschikt': 'M_lk_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_0; M_nws_2; M_nws_3',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'M_nws_4; S_nws_0; S_nws_2; S_nws_3'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': 'M_nws_3',
  'Gewenste placeholder': 'M_nws_0; M_nws_2;',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_0; S_nws_2;'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'M_nws_4',
  'Gewenste placeholder': 'M_nws_2; M_nws_3;',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_nws_0',
  'Vierde keus placeholder': 'S_nws_2; S_nws_3;'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'M_nws_4',
  'Gewenste placeholder': 'M_nws_2; M_nws_3;',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_nws_0',
  'Vierde keus placeholder': 'S_nws_2; S_nws_3; S_nws_0;'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': 'M_nws_2',
  'Gewenste placeholder': 'M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_nws_3;',
  'Vierde keus placeholder': 'S_nws_4; S_nws_5; S_nws_3'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_3; M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_nws_2',
  'Vierde keus placeholder': 'S_nws_3; S_nws_4; S_nws_5'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_2; M_nws_3; M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_2; S_nws_3; S_nws_4; S_nws_5'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'M_nws_0',
  'Gewenste placeholder': 'M_nws_2; M_nws_3; M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_2; S_nws_3; S_nws_4; S_nws_5'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'M_nws_3',
  'Vierde keus placeholder': 'M_nws_2; S_nws_4; S_nws_5'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'M_nws_3',
  'Vierde keus placeholder': 'M_nws_2; S_nws_4; S_nws_5'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_0; M_nws_2; M_nws_3; M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'M_nws_0; M_nws_2; M_nws_3; M_nws_4; M_nws_5',
  'Placeholder bij enigszins geschikt': 'M_nws_0; M_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'M_nws_2',
  'Gewenste placeholder': 'M_nws_0',
  'Placeholder bij enigszins geschikt': 'M_nws_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_0'},
 {'Artikelsoort': 'M_nws',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': 'M_nws_2',
  'Gewenste placeholder': 'M_nws_0',
  'Placeholder bij enigszins geschikt': 'M_nws_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_0'},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0; S_lk_2;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_lk_4'},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0; S_lk_2;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'S_lk_4',
  'Gewenste placeholder': 'S_lk_2;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_lk_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'S_lk_4',
  'Gewenste placeholder': 'S_lk_2;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_lk_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_lk_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_lk_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_2; S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_lk_0'},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'S_lk_0',
  'Gewenste placeholder': 'S_lk_2; S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_lk_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_lk_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0; S_lk_2; S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0; S_lk_2; S_lk_4;',
  'Placeholder bij enigszins geschikt': 'S_lk_0; S_lk_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0',
  'Placeholder bij enigszins geschikt': 'S_lk_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_lk_2'},
 {'Artikelsoort': 'S_lk',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_lk_0',
  'Placeholder bij enigszins geschikt': 'S_lk_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_lk_2'},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0; S_nws_2;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_4'},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0; S_nws_2;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'S_nws_4',
  'Gewenste placeholder': 'S_nws_2;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_nws_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': 'S_nws_4',
  'Gewenste placeholder': 'S_nws_2;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_nws_0',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_nws_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_nws_2',
  'Vierde keus placeholder': 'S_nws_0'},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_2; S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': 'S_nws_0',
  'Gewenste placeholder': 'S_nws_2; S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_2',
  'Top 8': 'Ja',
  'Tweede keus placeholder': 'S_nws_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_2',
  'Top 8': 'Nee',
  'Tweede keus placeholder': 'S_nws_2',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0; S_nws_2; S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0; S_nws_2; S_nws_4;',
  'Placeholder bij enigszins geschikt': 'S_nws_0; S_nws_2;',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0',
  'Placeholder bij enigszins geschikt': 'S_nws_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_2'},
 {'Artikelsoort': 'S_nws',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'S_nws_0',
  'Placeholder bij enigszins geschikt': 'S_nws_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'S_nws_2'},
 {'Artikelsoort': 'XS',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': '(geen waarde ingevuld)',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': 'XS_4'},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend of bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Dragend en bijplaat',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_4',
  'Placeholder bij enigszins geschikt': 'XS_4',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0; XS_4',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Flexibel',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0; XS_4',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Ja',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''},
 {'Artikelsoort': 'XS',
  'Beeld voor print': 'Ongeschikt',
  'Derde keus placeholder': '',
  'Gewenste placeholder': 'XS_0',
  'Placeholder bij enigszins geschikt': 'XS_0',
  'Top 8': 'Nee',
  'Tweede keus placeholder': '',
  'Vierde keus placeholder': ''}]



def is_valid_colname(c):
    if c is None:
        return False
    if isinstance(c, float) and math.isnan(c):
        return False
    s = str(c).replace("\u00a0", " ").strip()
    if s == "" or s.lower().startswith("unnamed"):
        return False
    return True

def normalize_colnames(cols):
    return [str(c).replace("\u00a0", " ").strip() for c in cols]

def drop_rows_type_verhaal_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Type verhaal" not in df.columns:
        return df
    tv = df["Type verhaal"].astype("string").str.replace("\u00a0", " ", regex=False).str.strip()
    before = len(df)
    df = df.loc[~tv.eq("Column")].copy()
    dropped = before - len(df)
    if dropped:
        print(f"Rijen verwijderd omdat Type verhaal='Column': {dropped}")
    return df.reset_index(drop=True)

def add_mapping_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # v18: maak Artikelsoort-logic robuust: accepteer 'Text length' óf 'Karakters'
    if "Text length" in df.columns and "Karakters" not in df.columns:
        df = df.rename(columns={"Text length": "Karakters"})

    # Artikelsoort
    # Artikelsoort (op basis van Karakters)
    if "Karakters" in df.columns:
        tl = pd.to_numeric(df["Karakters"], errors="coerce")
        tv = df["Type verhaal"].astype("string") if "Type verhaal" in df.columns else pd.Series(pd.NA, index=df.index)
        df["Artikelsoort"] = pd.NA
        df.loc[tl.eq(7200), "Artikelsoort"] = "XXL"
        df.loc[tl.eq(5400), "Artikelsoort"] = "XL"
        df.loc[tl.eq(4000), "Artikelsoort"] = "L"
        df.loc[tl.eq(2800) & (tv.eq("Lees") | tv.eq("Duiding")), "Artikelsoort"] = "M_lk"
        df.loc[tl.eq(2800) & tv.eq("Nieuws"), "Artikelsoort"] = "M_nws"
        df.loc[tl.eq(1800) & (tv.eq("Lees") | tv.eq("Duiding")), "Artikelsoort"] = "S_lk"
        df.loc[tl.eq(1800) & tv.eq("Nieuws"), "Artikelsoort"] = "S_nws"
        df.loc[tl.eq(1000), "Artikelsoort"] = "XS"
    else:
        df["Artikelsoort"] = pd.NA


    # Group → Leverancier
    if "Group" in df.columns:
        df["Leverancier"] = df["Group"].astype("string").map(GROUP_TO_LEVERANCIER)
        df = df.drop(columns=["Group"])
    else:
        df["Leverancier"] = pd.NA
        print("Waarschuwing: kolom 'Group' niet gevonden; 'Leverancier' blijft leeg.")

    # Auteur
    if "Assignee last name" in df.columns and "Assignee first name" in df.columns:
        lastn = df["Assignee last name"].astype("string")
        firstn = df["Assignee first name"].astype("string")
        df["Auteur"] = (lastn.fillna("") + ", " + firstn.fillna(""))
        df.loc[df["Auteur"].str.strip().isin([",", ", ", ""]), "Auteur"] = pd.NA
    else:
        df["Auteur"] = pd.NA
        print("Waarschuwing: assignee-kolommen ontbreken; 'Auteur' blijft leeg.")

    return df

def parse_excel_story_list(input_path: str) -> pd.DataFrame:
    raw = pd.read_excel(input_path, sheet_name="Story List", header=None, engine="openpyxl")
    if raw.shape[0] <= 5:
        raise ValueError("Te weinig rijen om headers te bepalen.")
    trimmed = raw.iloc[5:].reset_index(drop=True)
    header = trimmed.iloc[0].tolist()
    df = trimmed.iloc[1:].copy()
    df.columns = header

    df = df.loc[:, [c for c in df.columns if is_valid_colname(c)]]
    df.columns = normalize_colnames(df.columns)

    # v17: behoud Karakters, hernoem naar Karakters
    if "Karakters" in df.columns:
        df = df.rename(columns={"Karakters": "Karakters"})

    # v14+: kolomnaam vervangen
    if "Description" in df.columns and "Naam productie" not in df.columns:
        df = df.rename(columns={"Description": "Naam productie"})

    df = df.drop(columns=[c for c in COLUMNS_TO_DROP if c in df.columns])
    df = df.dropna(how="all").reset_index(drop=True)

    df = drop_rows_type_verhaal_column(df)
    df = add_mapping_columns(df)
    df = apply_placeholder_rules(df)
    df = df.drop(columns=[c for c in FINAL_COLUMNS_TO_DROP if c in df.columns])

    if "Description" in df.columns and "Naam productie" not in df.columns:
        df = df.rename(columns={"Description": "Naam productie"})

    return df

def save_df_single(df: pd.DataFrame) -> str:
    out = "parsed_story_list.xlsx"
    df.to_excel(out, index=False)
    return out

def _norm_series(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="string")
    return df[col].astype("string").str.replace("\u00a0", " ", regex=False).str.strip()

def _contains_any(series: pd.Series, allowed: list[str]) -> pd.Series:
    allowed_set = {a.strip().lower() for a in allowed}
    def cell_match(x):
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return False
        s = str(x).replace("\u00a0", " ").strip()
        if s == "" or s.lower() == "nan":
            return False
        parts = [p.strip().lower() for p in re.split(r"[;]", s) if p.strip()]
        return any(p in allowed_set for p in parts)
    return series.apply(cell_match)

def _equals_ci(series: pd.Series, value: str) -> pd.Series:
    b = (
        series.astype("string")
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
        .str.lower()
        .eq(str(value).strip().lower())
    )
    return b.fillna(False)

def _equals_any_ci(series: pd.Series, values: list[str] | set[str] | tuple[str, ...]) -> pd.Series:
    vals = [str(v).strip().lower() for v in values]
    s = (
        series.astype("string")
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
        .str.lower()
    )
    return s.isin(vals).fillna(False)

def _b2i(b: pd.Series) -> pd.Series:
    return b.fillna(False).astype(int)

def voorkeur_penalty(voorkeur: pd.Series, lijstcode: str) -> pd.Series:
    """v16: IF Voorkeurspositie != Nee AND != lijstcode THEN -20."""
    return _b2i((~_equals_ci(voorkeur, "Nee")) & (~_equals_ci(voorkeur, lijstcode))) * (-20)



def voorkeur_penalty_allowed(voorkeur: pd.Series, allowed_codes: list[str] | set[str] | tuple[str, ...]) -> pd.Series:
    """Variant: IF Voorkeurspositie != Nee AND not in allowed_codes THEN -20."""
    return _b2i((~_equals_ci(voorkeur, "Nee")) & (~_equals_any_ci(voorkeur, allowed_codes))) * (-20)



def apply_placeholder_rules(df_in: pd.DataFrame) -> pd.DataFrame:
    """V26 uitbreiding: voeg 5 placeholder-kolommen toe en vul ze volgens hardcoded mappingregels."""
    df = df_in.copy()

    out_cols = [
        "Gewenste placeholder",
        "Tweede keus placeholder",
        "Derde keus placeholder",
        "Vierde keus placeholder",
        "Placeholder bij enigszins geschikt"]
    for c in out_cols:
        if c not in df.columns:
            df[c] = ""

    art = _norm_series(df, "Artikelsoort")
    beeld = _norm_series(df, "Beeld voor print")
    top8 = _norm_series(df, "Top 8")

    for rule in PLACEHOLDER_RULES:
        r_art = str(rule.get("Artikelsoort", "")).strip()
        r_beeld = str(rule.get("Beeld voor print", "")).strip()
        r_top8 = str(rule.get("Top 8", "")).strip()

        if not r_art:
            continue

        mask = _equals_ci(art, r_art)

        # Beeld voor print
        if r_beeld:
            if r_beeld.lower() == "(geen waarde ingevuld)":
                mask = mask & (beeld.isna() | beeld.eq(""))
            else:
                mask = mask & _equals_ci(beeld, r_beeld)

        # Top 8
        if r_top8:
            mask = mask & _equals_ci(top8, r_top8)

        if not mask.any():
            continue

        # Schrijf outputs (exact zoals in mappingregels; leeg blijft leeg)
        for out_c in out_cols:
            v = str(rule.get(out_c, "")).replace("\u00a0", " ").strip()
            if v != "":
                df.loc[mask, out_c] = v

    return df

def add_prioscore_focus(df_in: pd.DataFrame, focus_value: str, voorkeur_code: str, leverancier_code: str, apply_penalty: bool) -> pd.DataFrame:
    """Prioscore-regels voor NM/ZU lijsten (single focuswaarde + Limburg-breed)."""
    df = df_in.copy()
    focus_raw = _norm_series(df, "Focusregio")
    pub = _norm_series(df, "Publicatiedwang")
    heel = _norm_series(df, "Heel Limburg")
    lev = _norm_series(df, "Leverancier")
    voorkeur = _norm_series(df, "Voorkeurspositie")
    top8 = _norm_series(df, "Top 8")

    is_focus = _contains_any(focus_raw, [focus_value])
    is_lb = _contains_any(focus_raw, ["Limburg-breed"])

    score = pd.Series(0, index=df.index, dtype="int64")
    score += _b2i(is_focus) * 3
    score += _b2i(_equals_ci(voorkeur, voorkeur_code)) * 25
    score += _b2i(is_focus & _equals_ci(pub, "Ja")) * 3
    score += _b2i(is_focus & _equals_ci(pub, "Nee")) * (-2)
    score += _b2i(is_focus & _equals_ci(heel, "ongeschikt")) * 2
    score += _b2i(is_lb) * 2
    score += _b2i(_equals_ci(lev, leverancier_code)) * 1
    score += _b2i(is_focus & _equals_ci(top8, "Ja")) * 5

    if apply_penalty:
        score += voorkeur_penalty_allowed(voorkeur, {"ND-01","ND-02","ND-03"})

    df["Prioscore"] = score
    return df.sort_values(by="Prioscore", ascending=False, kind="mergesort").reset_index(drop=True)

def add_prioscore_nd(df_in: pd.DataFrame, voorkeur_code: str, apply_penalty: bool) -> pd.DataFrame:
    """Prioscore-regels voor ND-01 / ND-02 / ND-03."""
    df = df_in.copy()
    focus_raw = _norm_series(df, "Focusregio")
    pub = _norm_series(df, "Publicatiedwang")
    heel = _norm_series(df, "Heel Limburg")
    voorkeur = _norm_series(df, "Voorkeurspositie")
    top8 = _norm_series(df, "Top 8")

    is_lb = _contains_any(focus_raw, ["Limburg-breed"])

    score = pd.Series(0, index=df.index, dtype="int64")
    score += _b2i(is_lb) * 3
    score += _b2i(_equals_ci(voorkeur, voorkeur_code)) * 25
    score += _b2i(is_lb & _equals_ci(pub, "Ja")) * 3
    score += _b2i(_equals_ci(heel, "moet mee")) * 10
    score += _b2i(_equals_ci(heel, "geschikt")) * 8
    score += _b2i(_equals_ci(top8, "Ja")) * 5

    if apply_penalty:
        score += voorkeur_penalty_allowed(voorkeur, {"ND-01","ND-02","ND-03"})

    df["Prioscore"] = score
    return df.sort_values(by="Prioscore", ascending=False, kind="mergesort").reset_index(drop=True)

def add_prioscore_ov(df_in: pd.DataFrame, focus_values: list[str], voorkeur_code: str, leverancier_codes: list[str]) -> pd.DataFrame:
    """Prioscore-regels voor OV-lijsten (geen v16 penalty!)."""
    df = df_in.copy()
    focus_raw = _norm_series(df, "Focusregio")
    pub = _norm_series(df, "Publicatiedwang")
    heel = _norm_series(df, "Heel Limburg")
    lev = _norm_series(df, "Leverancier")
    voorkeur = _norm_series(df, "Voorkeurspositie")
    top8 = _norm_series(df, "Top 8")

    is_focus_any = _contains_any(focus_raw, focus_values)
    is_lb = _contains_any(focus_raw, ["Limburg-breed"])

    score = pd.Series(0, index=df.index, dtype="int64")
    score += _b2i(is_focus_any) * 3
    score += _b2i(_equals_ci(voorkeur, voorkeur_code)) * 25
    score += _b2i(is_focus_any & _equals_ci(pub, "Ja")) * 3
    score += _b2i(is_focus_any & _equals_ci(pub, "Nee")) * (-2)
    score += _b2i(is_focus_any & _equals_ci(pub, "Ja") & _equals_ci(heel, "enigszins geschikt")) * 10
    score += _b2i(is_focus_any & _equals_ci(pub, "Ja") & _equals_ci(heel, "ongeschikt")) * 25
    score += _b2i(is_lb) * 2

    lev_any = None
    for code in leverancier_codes:
        cond = _equals_ci(lev, code)
        lev_any = cond if lev_any is None else (lev_any | cond)
    if lev_any is None:
        lev_any = pd.Series(False, index=df.index)
    score += _b2i(lev_any) * 1

    score += _b2i(is_focus_any & _equals_ci(top8, "Ja")) * 5

    df["Prioscore"] = score
    return df.sort_values(by="Prioscore", ascending=False, kind="mergesort").reset_index(drop=True)

def build_verhalenaanbod(df: pd.DataFrame) -> dict:
    focus_raw = _norm_series(df, "Focusregio")
    heel = _norm_series(df, "Heel Limburg")

    def include_focus(values):
        return _contains_any(focus_raw, values)

    def include_heel(values):
        return heel.isin(values).fillna(False)

    def exclude_focus_and_heel(focus_values, heel_value):
        return ~(include_focus(focus_values) & _equals_ci(heel, heel_value))

    sheets = {}

    nm_no = df.loc[include_focus(["Noord", "Limburg-breed"])].copy()

    # v26: XXL/XL alleen als Focusregio expliciet 'Noord' bevat
    nm_no_focus = focus_raw.loc[nm_no.index]
    nm_no_art = _norm_series(nm_no, "Artikelsoort")
    nm_no_xl = _equals_ci(nm_no_art, "XXL") | _equals_ci(nm_no_art, "XL")
    nm_no_has = _contains_any(nm_no_focus, ["Noord"])
    nm_no = nm_no.loc[~((~nm_no_has) & nm_no_xl)].copy()
    sheets["NM-NO"] = add_prioscore_focus(nm_no, "Noord", "NM-NO", "rNO", apply_penalty=True)

    nm_mi = df.loc[include_focus(["Midden", "Limburg-breed"])].copy()

    # v26: XXL/XL alleen als Focusregio expliciet 'Midden' bevat
    nm_mi_focus = focus_raw.loc[nm_mi.index]
    nm_mi_art = _norm_series(nm_mi, "Artikelsoort")
    nm_mi_xl = _equals_ci(nm_mi_art, "XXL") | _equals_ci(nm_mi_art, "XL")
    nm_mi_has = _contains_any(nm_mi_focus, ["Midden"])
    nm_mi = nm_mi.loc[~((~nm_mi_has) & nm_mi_xl)].copy()
    sheets["NM-MI"] = add_prioscore_focus(nm_mi, "Midden", "NM-MI", "rMI", apply_penalty=True)

    nm_ov_base = df.loc[include_focus(["Midden", "Noord", "Limburg-breed"])].copy()
    # sheets["NM-OV"] = add_prioscore_ov(nm_ov_base, ["Noord", "Midden"], "NM-OV", ["rNO", "rMI"])  # verwijderd uit output

    zu_sg = df.loc[include_focus(["Sittard", "Limburg-breed"])].copy()

    # v26: XXL/XL alleen als Focusregio expliciet 'Sittard' bevat
    zu_sg_focus = focus_raw.loc[zu_sg.index]
    zu_sg_art = _norm_series(zu_sg, "Artikelsoort")
    zu_sg_xl = _equals_ci(zu_sg_art, "XXL") | _equals_ci(zu_sg_art, "XL")
    zu_sg_has = _contains_any(zu_sg_focus, ["Sittard"])
    zu_sg = zu_sg.loc[~((~zu_sg_has) & zu_sg_xl)].copy()
    sheets["ZU-SG"] = add_prioscore_focus(zu_sg, "Sittard", "ZU-SG", "rSG", apply_penalty=True)

    zu_mh = df.loc[include_focus(["Maastricht", "Limburg-breed"])].copy()

    # v26: XXL/XL alleen als Focusregio expliciet 'Maastricht' bevat
    zu_mh_focus = focus_raw.loc[zu_mh.index]
    zu_mh_art = _norm_series(zu_mh, "Artikelsoort")
    zu_mh_xl = _equals_ci(zu_mh_art, "XXL") | _equals_ci(zu_mh_art, "XL")
    zu_mh_has = _contains_any(zu_mh_focus, ["Maastricht"])
    zu_mh = zu_mh.loc[~((~zu_mh_has) & zu_mh_xl)].copy()
    sheets["ZU-MH"] = add_prioscore_focus(zu_mh, "Maastricht", "ZU-MH", "rMH", apply_penalty=True)

    zu_ps = df.loc[include_focus(["Parkstad", "Limburg-breed"])].copy()

    # v26: XXL/XL alleen als Focusregio expliciet 'Parkstad' bevat
    zu_ps_focus = focus_raw.loc[zu_ps.index]
    zu_ps_art = _norm_series(zu_ps, "Artikelsoort")
    zu_ps_xl = _equals_ci(zu_ps_art, "XXL") | _equals_ci(zu_ps_art, "XL")
    zu_ps_has = _contains_any(zu_ps_focus, ["Parkstad"])
    zu_ps = zu_ps.loc[~((~zu_ps_has) & zu_ps_xl)].copy()
    sheets["ZU-PS"] = add_prioscore_focus(zu_ps, "Parkstad", "ZU-PS", "rPS", apply_penalty=True)

    zu_ov_base = df.loc[include_focus(["Sittard", "Parkstad", "Maastricht", "Limburg-breed"])].copy()
    # sheets["ZU-OV"] = add_prioscore_ov(zu_ov_base, ["Sittard", "Maastricht", "Parkstad"], "ZU-OV", ["rNO", "rMI"])  # verwijderd uit output

    nd_mask = include_heel(["geschikt", "moet mee"])
    nd_base = df.loc[nd_mask].copy()
    sheets["ND-01"] = add_prioscore_nd(nd_base, "ND-01", apply_penalty=True)
    sheets["ND-02"] = add_prioscore_nd(nd_base, "ND-02", apply_penalty=True)
    sheets["ND-03"] = add_prioscore_nd(nd_base, "ND-03", apply_penalty=True)

    return sheets
def _norm_text(series: pd.Series) -> pd.Series:
    return series.astype("string").str.replace("\u00a0", " ", regex=False).str.strip()

def build_kandidaten_stats(sheets: dict) -> pd.DataFrame:
    """
    v25: extra tabblad met statistieken (alleen in verhalenaanbod-bestand).
    Rijen voor: NM-NO, NM-MI, ZU-SG, ZU-PS, ZU-MH, ND-01.
    Per rij: Tabblad, B_aantal, B_karakters, B_gem_lengte, B_fotos, B_fotoscore.
    """
    foto_values = {"dragend of bijplaat", "dragend en bijplaat", "dragend", "bijplaat", "flexibel"}
    rows = []

    def add_row(sheet_name: str, focus_value: str):
        if sheet_name not in sheets:
            return
        df = sheets[sheet_name].copy()

        focus = _norm_text(df.get("Focusregio", pd.Series(pd.NA, index=df.index)))
        # v25.2: B_fotos opnieuw berekenen o.b.v. 'Beeld voor print' (met fallback)
        beeld_col = "Beeld voor print" if "Beeld voor print" in df.columns else "Beeldformaat"
        beeld = _norm_text(df.get(beeld_col, pd.Series(pd.NA, index=df.index)))

        # Focusregio kan meerdere waarden hebben (bv 'Sittard; Midden')
        is_focus = _contains_any(focus, [focus_value]).fillna(False)

        b_aantal = int(is_focus.sum())

        if "Karakters" in df.columns:
            kar = pd.to_numeric(df["Karakters"], errors="coerce").fillna(0)
            b_karakters = int(kar[is_focus].sum())
            # v26: B_publicatiedwang = som Karakters, exclusief rijen waar Publicatiedwang = Nee
            if "Publicatiedwang" in df.columns:
                pub = _norm_text(df.get("Publicatiedwang", pd.Series(pd.NA, index=df.index))).str.lower()
                mask_pub = pub != "nee"
                b_publicatiedwang = int(kar[is_focus & mask_pub].sum())
            else:
                b_publicatiedwang = b_karakters

        else:
            b_karakters = 0
            b_publicatiedwang = 0

        is_foto = beeld.str.lower().isin(foto_values).fillna(False)
        b_fotos = int((is_focus & is_foto).sum())
        b_fotoscore = (b_fotos / b_aantal) if b_aantal else 0.0
        b_gem_lengte = (b_karakters / b_aantal) if b_aantal else 0.0
        complex_score = 0.0
        if b_aantal < 3:
            complex_score += 4
        if b_aantal > 6:
            complex_score += 1
        if b_karakters < 6000:
            complex_score += 6
        if b_karakters < 9000:
            complex_score += 3
        if b_karakters < 11000:
            complex_score += 2
        if b_karakters < 12000:
            complex_score += 1
        if b_karakters > 16000:
            complex_score += 2
        if b_gem_lengte < 2400:
            complex_score += 4
        if b_gem_lengte < 2600:
            complex_score += 2
        if b_gem_lengte > 3400:
            complex_score += 2
        if b_gem_lengte > 4500:
            complex_score += 4
        if b_fotos == 0:
            complex_score += 8
        if b_fotos == 1:
            complex_score += 6
        if b_fotos == 2:
            complex_score += 4
        if b_fotos > 4:
            complex_score += 2
        if b_fotoscore < 0.24:
            complex_score += 3
        if b_fotoscore > 0.80:
            complex_score += 2
        complexiteit = complex_score / 2.5


        rows.append({
            "Tabblad": sheet_name,
            "B_aantal": b_aantal,
            "B_karakters": b_karakters,
            "B_publicatiedwang": b_publicatiedwang,
            "B_gem_lengte": b_gem_lengte,
            "B_fotos": b_fotos,
            "B_fotoscore": b_fotoscore,
            "Complexiteit": complexiteit,
        })


    def add_total_row(sheet_name: str):
        if sheet_name not in sheets:
            return
        df = sheets[sheet_name].copy()

        # v25.9: Totale verhalenlijst - zelfde berekeningen als de overige tabbladen, maar zonder Focusregio-filter
        beeld_col = "Beeld voor print" if "Beeld voor print" in df.columns else "Beeldformaat"
        beeld = _norm_text(df.get(beeld_col, pd.Series(pd.NA, index=df.index)))

        is_focus = pd.Series(True, index=df.index)
        b_aantal = int(is_focus.sum())

        if "Karakters" in df.columns:
            kar = pd.to_numeric(df["Karakters"], errors="coerce").fillna(0)
            b_karakters = int(kar[is_focus].sum())
            # v26: B_publicatiedwang = som Karakters, exclusief rijen waar Publicatiedwang = Nee
            if "Publicatiedwang" in df.columns:
                pub = _norm_text(df.get("Publicatiedwang", pd.Series(pd.NA, index=df.index))).str.lower()
                mask_pub = pub != "nee"
                b_publicatiedwang = int(kar[is_focus & mask_pub].sum())
            else:
                b_publicatiedwang = b_karakters

        else:
            b_karakters = 0
            b_publicatiedwang = 0

        is_foto = beeld.str.lower().isin(foto_values).fillna(False)
        b_fotos = int((is_focus & is_foto).sum())
        b_fotoscore = (b_fotos / b_aantal) if b_aantal else 0.0
        b_gem_lengte = (b_karakters / b_aantal) if b_aantal else 0.0
        complex_score = 0.0
        if b_aantal < 3:
            complex_score += 4
        if b_aantal > 6:
            complex_score += 1
        if b_karakters < 6000:
            complex_score += 6
        if b_karakters < 9000:
            complex_score += 3
        if b_karakters < 11000:
            complex_score += 2
        if b_karakters < 12000:
            complex_score += 1
        if b_karakters > 16000:
            complex_score += 2
        if b_gem_lengte < 2400:
            complex_score += 4
        if b_gem_lengte < 2600:
            complex_score += 2
        if b_gem_lengte > 3400:
            complex_score += 2
        if b_gem_lengte > 4500:
            complex_score += 4
        if b_fotos == 0:
            complex_score += 8
        if b_fotos == 1:
            complex_score += 6
        if b_fotos == 2:
            complex_score += 4
        if b_fotos > 4:
            complex_score += 2
        if b_fotoscore < 0.24:
            complex_score += 3
        if b_fotoscore > 0.80:
            complex_score += 2
        complexiteit = complex_score / 2.5

        rows.append({
            "Tabblad": sheet_name,
            "B_aantal": b_aantal,
            "B_karakters": b_karakters,
            "B_publicatiedwang": b_publicatiedwang,
            "B_gem_lengte": b_gem_lengte,
            "B_fotos": b_fotos,
            "B_fotoscore": b_fotoscore,
            "Complexiteit": complexiteit,
        })

    add_total_row("Totale verhalenlijst")

    add_row("NM-NO", "Noord")
    add_row("NM-MI", "Midden")
    add_row("ZU-SG", "Sittard")
    add_row("ZU-PS", "Parkstad")
    add_row("ZU-MH", "Maastricht")
    add_row("ND-01", "Limburg-breed")
    return pd.DataFrame(rows, columns=["Tabblad","B_aantal","B_karakters","B_publicatiedwang","B_gem_lengte","B_fotos","B_fotoscore","Complexiteit"])

def save_verhalenaanbod_xlsx(df: pd.DataFrame) -> str:
    out = "verhalenaanbod.xlsx"
    sheets = build_verhalenaanbod(df)
    # v25.10: zorg dat Stats ook een rij kan maken voor 'Totale verhalenlijst'
    # (deze sheet wordt apart weggeschreven, maar moet wel in de stats-berekening zitten)
    sheets["Totale verhalenlijst"] = df.copy()
    stats_df = build_kandidaten_stats(sheets)
    # Toegevoegde kolom Akpp_ondergrens
    def _safe_insert(_df, _loc, _col, _val):
        _loc = int(_loc)
        if _col in _df.columns:
            _df.drop(columns=[_col], inplace=True)
        _loc = max(0, min(_loc, _df.shape[1]))
        _df.insert(_loc, _col, _val)
    _ukpp = (stats_df['B_karakters'] * 0.7) / 2
    _ukpp = _ukpp.where(_ukpp < 4500, 4500)
    _safe_insert(stats_df, 8, 'Akpp_ondergrens', _ukpp)
    # Toegevoegde kolommen Akpp_range_normaal en Akpp_range_papierschaarste
    # Aangepaste kolom Akpp_range_verhalenschaarste (nieuwe berekening)
    _verhaal_min = (stats_df['Akpp_ondergrens'] / 1.4).round(0)
    _verhaal_max = (stats_df['Akpp_ondergrens'] * 1.37).round(0)
    _verhaal_min = _verhaal_min.where(_verhaal_min > 2700, 2700)
    _verhaal_max = _verhaal_max.where(_verhaal_max > 5500, 5500)
    _safe_insert(stats_df, 11, 'Akpp_range_verhalenschaarste', _verhaal_min.astype(int).astype(str) + ':' + _verhaal_max.astype(int).astype(str))

    _safe_insert(stats_df, 14, 'Akpp_range_verhalenschaarste_extra', _verhaal_min.astype(int).astype(str) + ':' + (_verhaal_max - 300).astype(int).astype(str))
    _safe_insert(stats_df, 9, 'Akpp_range_normaal', stats_df['Akpp_ondergrens'].astype(int).astype(str) + ':7200')
    _safe_insert(stats_df, 10, 'Akpp_range_papierschaarste', (stats_df['Akpp_ondergrens'] + 300).astype(int).astype(str) + ':7800')

    _safe_insert(stats_df, 12, 'Akpp_range_normaal_extra', (stats_df['Akpp_ondergrens'] + 500).astype(int).astype(str) + ':7200')
    _safe_insert(stats_df, 13, 'Akpp_range_papierschaarste_extra', (stats_df['Akpp_ondergrens'] + 1000).astype(int).astype(str) + ':7800')
    # Aangepaste kolom Akpp_range_verhalenschaarste (nieuwe berekening)
    _verhaal_min = (stats_df['Akpp_ondergrens'] / 1.4).round(0)
    _verhaal_max = (stats_df['Akpp_ondergrens'] * 1.37).round(0)
    _verhaal_min = _verhaal_min.where(_verhaal_min > 2700, 2700)
    _verhaal_max = _verhaal_max.where(_verhaal_max > 5500, 5500)
    _safe_insert(stats_df, 11, 'Akpp_range_verhalenschaarste', _verhaal_min.astype(int).astype(str) + ':' + _verhaal_max.astype(int).astype(str))
    for k in list(sheets.keys()):
        if "Description" in sheets[k].columns and "Naam productie" not in sheets[k].columns:
            sheets[k] = sheets[k].rename(columns={"Description": "Naam productie"})
    order = [
        "NM-NO","NM-MI","ZU-SG","ZU-MH","ZU-PS",
        "ND-01","ND-02","ND-03"
    ]

    # Voeg extra kolommen toe aan elk tabblad (behalve Logfile en Stats)
    _
    # --- vNEW: voeg kolom 'Classificatie' toe (kolom E) ---
    def _insert_classificatie(_df, _sheet_name):
        if _sheet_name in ["Logfile", "Stats", "Planningsvolgorde"]:
            return _df
        # bepaal classificatie o.b.v. Focusregio per tabblad
        _special = {
            "NM-NO": "Noord",
            "NM-MI": "Midden",
            "ZU-SG": "Sittard",
            "ZU-MH": "Maastricht",
            "ZU-PS": "Parkstad",
        }
        if "Classificatie" in _df.columns:
            _df = _df.drop(columns=["Classificatie"])
        if _sheet_name in _special and "Focusregio" in _df.columns:
            _term = _special[_sheet_name]
            _has = _df["Focusregio"].fillna("").astype(str).str.contains(_term, case=False, na=False)
            _vals = _has.map(lambda x: "A-keus; B-keus; C-keus;" if x else "B-keus; C-keus;")
        else:
            _vals = "A-keus; B-keus; C-keus;"
        # Kolom E = index 4 (0-based)
        _pos = 4 if len(_df.columns) >= 4 else len(_df.columns)
        _df.insert(_pos, "Classificatie", _vals)
        return _df

    sheets = {k: _insert_classificatie(v, k) for k, v in sheets.items()}
    extra_cols = ["Gekozen template", "Gekozen placeholder", "Plaatsing"]
    for _sheet_name, _df in sheets.items():
        if _sheet_name not in ["Logfile", "Stats", "Planningsvolgorde"]:
            for _c in extra_cols:
                if _c not in _df.columns:
                    _df[_c] = ""


    # --- vNEW: bewaar waarden voor 'Vierde keus placeholder' en schrijf ze later op kolom T ---
    _vierde_map = {}
    # Totale verhalenlijst krijgt dezelfde rijvolgorde als df
    if "Vierde keus placeholder" in df.columns:
        _vierde_map["Totale verhalenlijst"] = df["Vierde keus placeholder"].fillna("").astype(str).tolist()

    for _sheet_name, _df in sheets.items():
        if _sheet_name not in ["Logfile", "Stats", "Planningsvolgorde"] and "Vierde keus placeholder" in _df.columns:
            _vierde_map[_sheet_name] = _df["Vierde keus placeholder"].fillna("").astype(str).tolist()
            sheets[_sheet_name] = _df.drop(columns=["Vierde keus placeholder"])

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Extra tabblad: Totale verhalenlijst (afkomstig van knop 3)
        _template_cols = sheets[order[0]].columns.tolist()
        total_df = df.copy()
        # vNEW: Classificatie ook in 'Totale verhalenlijst' (altijd A-keus; B-keus; C-keus;)
        if "Classificatie" not in total_df.columns:
            _pos = 4 if len(total_df.columns) >= 4 else len(total_df.columns)
            total_df.insert(_pos, "Classificatie", "A-keus; B-keus; C-keus;")
        if "Description" in total_df.columns and "Naam productie" not in total_df.columns:
            total_df = total_df.rename(columns={"Description": "Naam productie"})
        total_df = total_df.reindex(columns=_template_cols)
        total_df.to_excel(writer, sheet_name="Totale verhalenlijst", index=False)

        for name in order:
            sheets[name].to_excel(writer, sheet_name=name, index=False)
            # (versimpeld) Geen extra P2_/P3_-kopieën meer

        # Extra tabblad: Logfile (leeg, alleen kolommen)
        pd.DataFrame(columns=["Timestamp", "Beschrijving"]).to_excel(
            writer, sheet_name="Logfile", index=False
        )

        # --- vNEW: fixed AKPP ranges voor specifieke tabbladen (overschrijft bestaande waarden) ---
        _fixed_stats_rows = [
            {"Tabblad":"ND-01","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"NM-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"ZU-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"ND-02","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"ND-03","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P2_ND-01","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P2_NM-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P2_ZU-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P2_ND-02","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P2_ND-03","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P3_ND-01","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P3_NM-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P3_ZU-OV","Akpp_range_normaal":"4500:7200","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"3214:6165","Akpp_range_normaal_extra":"4800:7200","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P3_ND-02","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
            {"Tabblad":"P3_ND-03","Akpp_range_normaal":"3800:7400","Akpp_range_papierschaarste":"4800:7800","Akpp_range_verhalenschaarste":"2900:6165","Akpp_range_normaal_extra":"4500:7400","Akpp_range_papierschaarste_extra":"5500:7800","Akpp_range_verhalenschaarste_extra":"3214:5865"},
        ]
        _fixed_blank_cols = ["B_aantal","B_karakters","B_publicatiedwang","B_gem_lengte","B_fotos","B_fotoscore","Complexiteit","Akpp_ondergrens"]
        # zorg dat kolommen bestaan
        for _c in _fixed_blank_cols + [
            "Akpp_range_normaal","Akpp_range_papierschaarste","Akpp_range_verhalenschaarste",
            "Akpp_range_normaal_extra","Akpp_range_papierschaarste_extra","Akpp_range_verhalenschaarste_extra"
        ]:
            if _c not in stats_df.columns:
                stats_df[_c] = pd.NA
        for _row in _fixed_stats_rows:
            _tb = _row["Tabblad"]
            _mask = (stats_df["Tabblad"] == _tb) if "Tabblad" in stats_df.columns else pd.Series([False]*len(stats_df))
            if _mask.any():
                _i = stats_df.index[_mask][0]
                for _c in _fixed_blank_cols:
                    stats_df.at[_i, _c] = pd.NA
                for _c,_v in _row.items():
                    stats_df.at[_i, _c] = _v
            else:
                _new = {c: pd.NA for c in stats_df.columns}
                for _c in _fixed_blank_cols:
                    _new[_c] = pd.NA
                for _c,_v in _row.items():
                    _new[_c] = _v
                stats_df = pd.concat([stats_df, pd.DataFrame([_new])], ignore_index=True)

        # Extra tabblad met statistieken
        # vSIMPLIFY2: Stats uitbreiden met ND-02 en ND-03 op basis van ND-01
        # Regel:
        # 1) Verwijder alle rijen beneden 'ND-01'
        # 2) Kopieer vervolgens 'ND-01' twee keer, hernoem naar 'ND-02' en 'ND-03'
        # 3) Plaats deze direct onder 'ND-01'
        if not stats_df.empty:
            _key_col = "Tabblad" if "Tabblad" in stats_df.columns else stats_df.columns[0]
            _nd01_idx = stats_df.index[
                stats_df[_key_col].astype(str).str.replace("\u00a0", " ", regex=False).str.strip().str.upper() == "ND-01"
            ]
            if len(_nd01_idx) > 0:
                _nd01_pos = int(_nd01_idx[0])

                # Alles onder ND-01 verwijderen
                stats_df = stats_df.iloc[: _nd01_pos + 1].copy()

                # ND-01 rij dupliceren naar ND-02 / ND-03
                _nd01_row = stats_df.iloc[_nd01_pos].copy()
                _nd02_row = _nd01_row.copy()
                _nd03_row = _nd01_row.copy()
                _nd02_row[_key_col] = "ND-02"
                _nd03_row[_key_col] = "ND-03"

                stats_df = pd.concat(
                    [stats_df, pd.DataFrame([_nd02_row, _nd03_row])],
                    ignore_index=True,
                )

        stats_df.to_excel(writer, sheet_name="Stats", index=False)

        # Extra tabblad met planningsvolgorde
        _plan_order_keys = ["NM-NO","NM-MI","ZU-SG","ZU-MH","ZU-PS"]
        _sorted = (
            stats_df[stats_df["Tabblad"].isin(_plan_order_keys)]
            .sort_values("Complexiteit", ascending=False)["Tabblad"]
            .tolist()
        )
        plan_rows = ["Planningsvolgorde"] + _sorted
        plan_df = pd.DataFrame(plan_rows)
        plan_df.to_excel(writer, sheet_name="Planningsvolgorde", index=False, header=False)
    # --- vNEW: voeg 'Vierde keus placeholder' kolom toe + vaste teksten (AA1/AB1) ---
    # Belangrijk: we openen de workbook één keer, passen alles toe, en saven één keer.
    try:
        from openpyxl import load_workbook
        from copy import copy as _copy

        _wb2 = load_workbook(out)

        # 1) Vierde keus placeholder kolom op T (20) op alle relevante tabbladen
        _exclude = {"logfile", "stats", "planningsvolgorde"}
        for _ws in _wb2.worksheets:
            if _ws.title.strip().lower() in _exclude:
                continue

            # Insert een nieuwe kolom op T (20)
            _ws.insert_cols(20)

            # Header (rij 1) in de nieuwe kolom
            _ws.cell(row=1, column=20).value = "Vierde keus placeholder"

            # Neem dezelfde opmaak over als de andere headers (kopieer van kolom 19, rij 1)
            try:
                _src = _ws.cell(row=1, column=19)
                _dst = _ws.cell(row=1, column=20)
                _dst.font = _copy(_src.font)
                _dst.border = _copy(_src.border)
                _dst.fill = _copy(_src.fill)
                _dst.number_format = _src.number_format
                _dst.protection = _copy(_src.protection)
                _dst.alignment = _copy(_src.alignment)
            except Exception:
                pass

            # 2) Vul kolom T met de eerder berekende waarden (zelfde rijvolgorde als in DataFrame)
            try:
                _vals = _vierde_map.get(_ws.title, None)
                if (_vals is None or len(_vals) == 0) and (str(_ws.title).startswith("P2_") or str(_ws.title).startswith("P3_")):
                    _base = str(_ws.title).split("_", 1)[1]
                    _vals = _vierde_map.get(_base, [])
                if _vals is None:
                    _vals = []

                # data start op rij 2
                for _i, _v in enumerate(_vals, start=2):
                    _ws.cell(row=_i, column=20).value = _v
            except Exception:
                pass


        # 2b) Placeholder-concessie header in kolom Z (26) op alle relevante tabbladen
        for _ws in _wb2.worksheets:
            if _ws.title.strip().lower() in _exclude:
                continue

            _ws.cell(row=1, column=26).value = "Placeholder-concessie"

            # Neem dezelfde opmaak over als de andere headers (kopieer van kolom Y, rij 1)
            try:
                _src = _ws.cell(row=1, column=25 if _ws.max_column >= 25 else 1)
                _dst = _ws.cell(row=1, column=26)
                _dst.font = _copy(_src.font)
                _dst.border = _copy(_src.border)
                _dst.fill = _copy(_src.fill)
                _dst.number_format = _src.number_format
                _dst.protection = _copy(_src.protection)
                _dst.alignment = _copy(_src.alignment)
            except Exception:
                pass


        # 3) Vaste teksten per tabblad (AA1 en AB1)
        _AA1 = {
            "NM-OV": "Totale verhalenlijst",
            "ZU-OV": "Totale verhalenlijst",
            "ND-01": "Totale verhalenlijst",
            "ND-02": "Totale verhalenlijst",
            "ND-03": "Totale verhalenlijst",
            "P2_NM-OV": "P2_Totale verhalenlijst",
            "P2_ZU-OV": "P2_Totale verhalenlijst",
            "P2_ND-01": "P2_Totale verhalenlijst",
            "P2_ND-02": "P2_Totale verhalenlijst",
            "P2_ND-03": "P2_Totale verhalenlijst",
            "P3_NM-OV": "P3_Totale verhalenlijst",
            "P3_ZU-OV": "P3_Totale verhalenlijst",
            "P3_ND-01": "P3_Totale verhalenlijst",
            "P3_ND-02": "P3_Totale verhalenlijst",
            "P3_ND-03": "P3_Totale verhalenlijst",
            "NM-NO": "Totale verhalenlijst",
            "NM-MI": "Totale verhalenlijst",
            "ZU-SG": "Totale verhalenlijst",
            "ZU-MH": "Totale verhalenlijst",
            "ZU-PS": "Totale verhalenlijst",
        }
        _AB1 = {
            "NM-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; ZU-OV; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "ZU-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "ND-01": "Totale verhalenlijst",
            "ND-02": "Totale verhalenlijst",
            "ND-03": "Totale verhalenlijst",
            "P2_NM-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; P2_ZU-OV; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "P2_ZU-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; P2_NM-OV; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "P2_ND-01": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "P2_ND-02": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "P2_ND-03": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P3_NM-OV; P3_ZU-OV; P3_ND-01; P3_ND-02; P3_ND-03;",
            "P3_NM-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; P3_ZU-OV; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03;",
            "P3_ZU-OV": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; P3_NM-OV; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03;",
            "P3_ND-01": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03;",
            "P3_ND-02": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03;",
            "P3_ND-03": "Totale verhalenlijst; P2_Totale verhalenlijst; P3_Totale verhalenlijst; NM-OV; ZU-OV; ND-01; ND-02; ND-03; P2_NM-OV; P2_ZU-OV; P2_ND-01; P2_ND-02; P2_ND-03;",
            "NM-NO": "Totale verhalenlijst",
            "NM-MI": "Totale verhalenlijst",
            "ZU-SG": "Totale verhalenlijst",
            "ZU-MH": "Totale verhalenlijst",
            "ZU-PS": "Totale verhalenlijst",
        }

        for _sh, _val in _AA1.items():
            if _sh in _wb2.sheetnames:
                _wb2[_sh]["AA1"].value = _val

        for _sh, _val in _AB1.items():
            if _sh in _wb2.sheetnames:
                _wb2[_sh]["AB1"].value = _val

        _wb2.save(out)
    except Exception as _e:
        print("Waarschuwing: kon vaste teksten / placeholder-kolom niet (volledig) toepassen:", _e)


    
    # --- vNEW: herorder tabbladen volgens gewenste volgorde ---
    try:
        _desired = [
            "Stats",
            "Logfile",
            "Planningsvolgorde",
            "Totale verhalenlijst",
            "NM-NO",
            "NM-MI",
            "NM-OV",
            "ZU-SG",
            "ZU-MH",
            "ZU-PS",
            "ZU-OV",
            "ND-01",
            "ND-02",
            "ND-03", 
            
        ]

        _existing = list(_wb2.sheetnames)
        _suffix2 = [s for s in _existing if s.startswith("P2_")]
        _suffix3 = [s for s in _existing if s.startswith("P3_")]

        _ordered = [s for s in _desired if s in _existing]
        _ordered += [s for s in _suffix2 if s not in _ordered]
        _ordered += [s for s in _suffix3 if s not in _ordered]
        _ordered += [s for s in _existing if s not in _ordered]  # eventuele overige tabbladen achteraan

        _wb2._sheets = [_wb2[s] for s in _ordered]
    except Exception as _e2:
        print("Waarschuwing: kon tabbladen niet herordenen:", _e2)
    _wb2.save(out)

    return out

print("Klaar. Run de volgende cel.")

def verhoog_ondergrens(range_str, verhoging):
    """Verwacht formaat 'onder:boven' (bv. '4500:7200'). Verhoogt alleen de ondergrens."""
    if not range_str or ":" not in str(range_str):
        return ""
    try:
        onder, boven = str(range_str).split(":", 1)
        nieuwe_onder = int(str(onder).strip()) + int(verhoging)
        return f"{nieuwe_onder}:{str(boven).strip()}"
    except Exception:
        return ""

def verlaag_bovengrens(range_str, verlaging):
    """Verwacht formaat 'onder:boven' (bv. '3214:6165'). Verlaagt alleen de bovengrens."""
    if not range_str or ":" not in str(range_str):
        return ""
    try:
        onder, boven = str(range_str).split(":", 1)
        nieuwe_boven = int(str(boven).strip()) - int(verlaging)
        return f"{str(onder).strip()}:{nieuwe_boven}"
    except Exception:
        return ""

__all__ = ['parse_excel_story_list','save_verhalenaanbod_xlsx']
